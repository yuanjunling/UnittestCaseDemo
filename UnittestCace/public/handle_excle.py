#coding=utf-8
import openpyxl
from openpyxl import load_workbook


class HandExcel:
    def load_excel(self,url):
        # 加载exce
        open_excel = openpyxl.load_workbook(url)
        return open_excel
    def get_sheet_data(self,url,index=None):
        # 加载试验sheet的内容
        sheet_name = self.load_excel(url).sheetnames
        if index == None:
            index = 0
        data = self.load_excel(url)[sheet_name[index]]
        return  data

    def get_cell_value(self,url,row,cols,index=None):
        # 获取某一个单元格内容
        data= self.get_sheet_data(url,index).cell(row=row,column=cols).value
        return data
    def get_rows(self,url,index=None):
        #获取行数
        row = self.get_sheet_data(url,index).max_row
        return row
    def get_rows_value(self,url,row,index=None):
        #获取某一行的内容
        row_list = []
        for i in self.get_sheet_data(url,index)[row]:
            row_list.append(i.value)
        return row_list
        # 调用此方法的时候，写入一行数据到excel最后一行
    def write_cell_content(self,url,array,sheet=None):
        workbook1 = load_workbook(url)
        if sheet==None:
            sheet='Mysheet'
        sheet = workbook1[sheet]
        sheet.append(array)
        workbook1.save(url)
        return sheet.append(array)
handle=HandExcel()

